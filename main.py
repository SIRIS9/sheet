import asyncio
import logging
import json
import os
import time
import traceback
import urllib.parse
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import CommandStart
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, WebAppInfo, ReplyKeyboardRemove, FSInputFile
from aiogram.exceptions import TelegramAPIError

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

TOKEN = "8794859325:AAGRCQumN4To6ZEdYPkqS1d2O70XQZ1kovc"
WEBAPP_URL = "https://siris9.github.io/sheet/"
EXCEL_FILE = "users_data.xlsx"
ADMIN_ID = 320132920 

bot = Bot(token=TOKEN)
dp = Dispatcher()

logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ================= دوال الإكسل المتقدمة =================
def format_and_autofit_excel(ws):
    ws.sheet_view.rightToLeft = True
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=11)
    center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin_border = Border(
        left=Side(style='thin', color="BFBFBF"), right=Side(style='thin', color="BFBFBF"),
        top=Side(style='thin', color="BFBFBF"), bottom=Side(style='thin', color="BFBFBF")
    )

    for row_idx, row in enumerate(ws.iter_rows()):
        for cell in row:
            cell.alignment = center_aligned
            cell.border = thin_border
            if row_idx == 0:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.font = data_font

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 6)

def init_excel():
    registered = set()
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "قاعدة البيانات"
        ws.append(["معرف المستخدم (ID)", "الاسم الرباعي واللقب", "اسم الأم الثلاثي", "التحصيل الدراسي", "سنة التخرج", "رقم الهاتف"])
        format_and_autofit_excel(ws)
        wb.save(EXCEL_FILE)
    else:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                registered.add(str(row[0]))
    return registered

registered_users = init_excel()

def get_user_data(user_id):
    if not os.path.exists(EXCEL_FILE): return None
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(user_id):
            return {
                "fullName": str(row[1] or ""),
                "motherName": str(row[2] or ""),
                "education": str(row[3] or ""),
                "gradYear": str(row[4] or ""),
                "phone": str(row[5] or "")
            }
    return None

def save_or_update_excel(effective_user_id, data, is_admin_manual_add=False):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    action_type = ""

    if is_admin_manual_add:
        actual_id = f"Manual-{int(time.time())}"
        ws.append([actual_id, data.get("fullName"), data.get("motherName"), data.get("education"), data.get("gradYear"), data.get("phone")])
        action_type = "admin_add"
    else:
        found_row = None
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=1).value) == str(effective_user_id):
                found_row = row
                break

        if found_row:
            ws.cell(row=found_row, column=2, value=data.get("fullName", ""))
            ws.cell(row=found_row, column=3, value=data.get("motherName", ""))
            ws.cell(row=found_row, column=4, value=data.get("education", ""))
            ws.cell(row=found_row, column=5, value=data.get("gradYear", ""))
            ws.cell(row=found_row, column=6, value=data.get("phone", ""))
            action_type = "update"
        else:
            ws.append([str(effective_user_id), data.get("fullName", ""), data.get("motherName", ""), data.get("education", ""), data.get("gradYear", ""), data.get("phone")])
            registered_users.add(str(effective_user_id))
            action_type = "new"

    format_and_autofit_excel(ws)
    wb.save(EXCEL_FILE)
    return action_type

# ================= دوال البوت =================

@dp.message(CommandStart())
async def command_start_handler(message: types.Message) -> None:
    user_id = str(message.from_user.id)
    user_name = message.from_user.first_name
    
    if user_id == str(ADMIN_ID):
        buttons = [[KeyboardButton(text="👨‍💻 إضافة مستخدم جديد (يدوي)", web_app=WebAppInfo(url=f"{WEBAPP_URL}?admin_add=true"))]]
        if user_id in registered_users:
            user_data = get_user_data(user_id)
            if user_data:
                query_params = urllib.parse.urlencode(user_data)
                buttons.append([KeyboardButton(text="🔄 تحديث بياناتي الشخصية", web_app=WebAppInfo(url=f"{WEBAPP_URL}?{query_params}"))])

        keyboard = ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)
        await message.answer(
            f"أهلاً بك حضرة المدير 👨‍💻\n\n"
            f"**الخيارات المتاحة لك:**\n"
            f"- استخدم الزر أدناه لإضافة سجلات جديدة.\n"
            f"- **لتعديل بيانات أي مستخدم:** فقط أرسل لي الـ ID الخاص به في رسالة (مثال: `7140294224`).",
            reply_markup=keyboard,
            parse_mode="Markdown"
        )
        return

    if user_id in registered_users:
        user_data = get_user_data(user_id)
        if user_data:
            query_params = urllib.parse.urlencode(user_data)
            kb = [[KeyboardButton(text="🔄 تحديث استمارتي", web_app=WebAppInfo(url=f"{WEBAPP_URL}?{query_params}"))]]
            keyboard = ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)
            await message.answer(f"أهلاً بعودتك {user_name} 🌟\nلقد تم توثيق بياناتك مسبقاً. لتحديثها، اضغط على الزر أدناه.", reply_markup=keyboard)
        return

    kb = [[KeyboardButton(text="📝 فتح الاستمارة الإلكترونية", web_app=WebAppInfo(url=WEBAPP_URL))]]
    keyboard = ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)
    await message.answer(f"أهلاً وسهلاً {user_name}.\nالرقم التعريفي الخاص بك: `{user_id}`\n\nيرجى الضغط أدناه لتسجيل بياناتك.", reply_markup=keyboard, parse_mode="Markdown")

# 🔵 [جديد] التقاط الـ ID لتعديل مستخدم من قبل المدير
@dp.message(lambda msg: msg.text and msg.text.isdigit() and str(msg.from_user.id) == str(ADMIN_ID))
async def admin_edit_user_handler(message: types.Message):
    target_id = message.text.strip()
    user_data = get_user_data(target_id)
    
    if user_data:
        # إضافة target_id للبيانات المشفرة بالرابط
        user_data['target_id'] = target_id
        query_params = urllib.parse.urlencode(user_data)
        custom_url = f"{WEBAPP_URL}?{query_params}"
        
        kb = [[KeyboardButton(text=f"⚙️ تعديل حساب ({target_id})", web_app=WebAppInfo(url=custom_url))]]
        keyboard = ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)
        
        await message.answer(
            f"تم العثور على سجل للمستخدم `{target_id}` في النظام 🔍\n\n"
            f"👤 **الاسم:** {user_data['fullName']}\n"
            f"لتعديل بيانات هذا المستخدم، اضغط على الزر أدناه:",
            reply_markup=keyboard,
            parse_mode="Markdown"
        )
    else:
        await message.answer("❌ عذراً أيها المدير، لم يتم العثور على أي سجل بهذا الرقم التعريفي في الإكسل.")

@dp.message(F.web_app_data)
async def web_app_data_handler(message: types.Message):
    user_id = str(message.from_user.id)
    is_admin = (user_id == str(ADMIN_ID))

    try:
        data = json.loads(message.web_app_data.data)
        target_id = data.get("target_id", "")
        
        is_admin_editing_other = False
        effective_user_id = user_id

        # إذا كان المرسل هو المدير، ويوجد target_id في الاستمارة (يعني أنه يعدل بيانات شخص آخر)
        if is_admin and target_id and target_id != str(ADMIN_ID):
            effective_user_id = target_id
            is_admin_editing_other = True
            
        is_admin_manual_add = False
        if is_admin and message.web_app_data.button_text and "إضافة مستخدم جديد" in message.web_app_data.button_text:
            is_admin_manual_add = True

        action = save_or_update_excel(effective_user_id, data, is_admin_manual_add=is_admin_manual_add)
        
        if is_admin_editing_other:
            await message.answer(f"✅ تم تحديث بيانات المستخدم `{effective_user_id}` بنجاح.", reply_markup=ReplyKeyboardRemove(), parse_mode="Markdown")
        elif action == "admin_add":
            await message.answer("✅ تم إدراج السجل الجديد يدوياً بنجاح.", reply_markup=ReplyKeyboardRemove())
        elif action == "update":
            await message.answer("🔄 تم تحديث وتعديل بياناتك في قاعدة البيانات بنجاح.", reply_markup=ReplyKeyboardRemove())
        else:
            await message.answer("✅ تم استلام وتوثيق بياناتك بنجاح في قاعدة البيانات المركزية.\nشكراً لتعاونكم.", reply_markup=ReplyKeyboardRemove())
        
        # إشعار المدير
        if is_admin_editing_other:
            action_title = "⚙️ تحديث بيانات من قبل المدير"
        elif action == "new":
            action_title = "🟢 تسجيل جديد"
        elif action == "update":
            action_title = "🔄 تحديث بيانات مستخدم"
        else:
            action_title = "👨‍💻 إدخال يدوي من المدير"
            
        admin_notification = (
            f"{action_title}!\n\n"
            f"👤 **الاسم:** {data.get('fullName')}\n"
            f"🆔 **المعرف:** `{effective_user_id}`\n"
            f"📱 **الهاتف:** {data.get('phone')}\n\n"
            f"📁 أحدث نسخة من سجل البيانات في الأسفل 👇"
        )
        
        if not is_admin_manual_add and not is_admin_editing_other:
            await bot.send_message(ADMIN_ID, admin_notification, parse_mode="Markdown")
        else:
            # إذا قام المدير بنفسه بالعملية، نرسل له الملف فقط كنسخة احتياطية
            pass
        
        excel_document = FSInputFile(EXCEL_FILE)
        await bot.send_document(ADMIN_ID, document=excel_document, caption="ملف قاعدة البيانات المحدث")
        
    except Exception as e:
        logger.error(f"Error processing web app data: {e}")
        await message.answer("❌ حدث خطأ تقني أثناء معالجة البيانات. يرجى المحاولة لاحقاً.")

@dp.errors()
async def global_error_handler(event: types.ErrorEvent):
    logger.error("Exception caused by event %s", event.update, exc_info=event.exception)
    exc_info = traceback.format_exception(type(event.exception), event.exception, event.exception.__traceback__)
    exc_string = "".join(exc_info)[-3000:]
    error_msg = f"🔴 **تنبيه النظام: حدوث تحطم!**\n\n**نوع الخطأ:** `{type(event.exception).__name__}`\n**التفاصيل:**\n```python\n{exc_string}\n```"
    try:
        await bot.send_message(ADMIN_ID, error_msg, parse_mode="Markdown")
    except:
        pass

async def main() -> None:
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
